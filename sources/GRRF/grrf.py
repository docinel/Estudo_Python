from datetime import datetime
import pyautogui as pa
import timeit
import openpyxl as xl

start_time = timeit.default_timer()

wb = xl.load_workbook(r'Z:\011-TI\017-BI_TESTES\
                      BI_BASE_RELATORIOS\relatorio_auditor.xlsx')

ws = wb['GRRF_B']

qtd = int(pa.prompt(text='Digite a quantidade de GRRF:',
                    title='QTD LINHAS', default='')) + 2
#  ABRE O MENTOR
pa.hotkey('win', 'd')
pa.sleep(1)
pa.getWindowsWithTitle('Touch Comp ERP')[0].maximize()

# NOVO ITEM
pa.doubleClick(1220, 164), pa.press('del'), pa.click(1220, 164)
meu_dic = {}
pa.sleep(5)
for cont in range(2, qtd):
    meu_dic[f'A{cont}'] = ws[f'A{cont}'].value
    meu_dic[f'B{cont}'] = ws[f'B{cont}'].value
    meu_dic[f'C{cont}'] = ws[f'C{cont}'].value

for cont in range(2, qtd):
    pa.write(str(meu_dic[f'A{cont}']))
    pa.sleep(1)
    pa.press('enter')
    pa.sleep(1.5)
    pa.click(1157, 429)
    pa.sleep(1)
    data = ws[f'C{cont}'].value
    data_t = datetime.date(data)
    data_te = format(data_t, '%d%m%Y')
    data_t_ano = format(data_t, '%Y')
    pa.write(data_te)
    pa.press('tab')
    pa.sleep(1)
    pa.press('enter')
    pa.sleep(3)
    pa.press('left')
    pa.write(str(meu_dic[f'B{cont}']))
    pa.write('_' + data_t_ano + '_')
    pa.sleep(1.6)
    pa.press('enter', presses=3, interval=2)
    pa.doubleClick(1220, 164), pa.press('del'), pa.click(1220, 164)

print((timeit.default_timer() - start_time)/60)
