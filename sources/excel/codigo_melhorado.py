import openpyxl as xl

wb = xl.load_workbook(r'sources/excel/calculo.xlsx')
ws = wb.active

meu_dic = {}
for cont in range(2, 11):
    meu_dic[f'A{cont}'] = ws[f'A{cont}'].value
    meu_dic[f'B{cont}'] = ws[f'B{cont}'].value

for cont in range(2, 11):
    print(meu_dic[f'A{cont}'])
    print(meu_dic[f'B{cont}'])

wb.close()