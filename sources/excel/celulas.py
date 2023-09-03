import openpyxl as xl

wb = xl.load_workbook(r'sources/excel/calculo.xlsx')
ws = wb.active

meu_dic = {'A2': ws['A2'].value, 'B2': ws['B2'].value,
           'A3': ws['A3'].value, 'B3': ws['B3'].value,
           'A4': ws['A4'].value, 'B4': ws['B4'].value,
           'A5': ws['A5'].value, 'B5': ws['B5'].value,
           'A6': ws['A6'].value, 'B6': ws['B6'].value,
           'A7': ws['A7'].value, 'B7': ws['B7'].value,
           'A8': ws['A8'].value, 'B8': ws['B8'].value,
           'A9': ws['A9'].value, 'B9': ws['B9'].value,
           'A10': ws['A10'].value, 'B10': ws['B10'].value,
           'A11': ws['A11'].value, 'B11': ws['B11'].value,
           'A12': ws['A12'].value, 'B12': ws['B12'].value,
           'A13': ws['A13'].value, 'B13': ws['B13'].value,
           'A14': ws['A14'].value, 'B14': ws['B14'].value,
           'A15': ws['A15'].value, 'B15': ws['B15'].value,
           'A16': ws['A16'].value, 'B16': ws['B16'].value,
           'A17': ws['A17'].value, 'B17': ws['B17'].value,
           'A18': ws['A18'].value, 'B18': ws['B18'].value,
           'A19': ws['A19'].value, 'B19': ws['B19'].value,
           'A20': ws['A20'].value, 'B20': ws['B20'].value,
           'A21': ws['A21'].value, 'B21': ws['B21'].value,
           'A22': ws['A22'].value, 'B22': ws['B22'].value,
           'A23': ws['A23'].value, 'B23': ws['B23'].value,
           'A24': ws['A24'].value, 'B24': ws['B24'].value,
           }

cont = 2
while cont < 19:
    print(meu_dic[f'A{cont}'])
    print(meu_dic[f'B{cont}'])
    cont +=1   

# 
# 
# A5 = ws['A5'].value
# B5 = ws['B5'].value
# A6 = ws['A6'].value
# B6 = ws['B6'].value
# A7 = ws['A7'].value
# B7 = ws['B7'].value
# A8 = ws['A8'].value
# B8 = ws['B8'].value
# A9 = ws['A9'].value
# B9 = ws['B9'].value
# A10 = ws['A10'].value
# B10 = ws['B10'].value

# A2 = ws['A2'].value
# B2 = ws['B2'].value
# A3 = ws['A3'].value
# B3 = ws['B3'].value
# A4 = ws['A4'].value
# B4 = ws['B4'].value
# A5 = ws['A5'].value
# B5 = ws['B5'].value
# A6 = ws['A6'].value
# B6 = ws['B6'].value
# A7 = ws['A7'].value
# B7 = ws['B7'].value
# A8 = ws['A8'].value
# B8 = ws['B8'].value
# A9 = ws['A9'].value
# B9 = ws['B9'].value
# A10 = ws['A10'].value
# B10 = ws['B10'].value









