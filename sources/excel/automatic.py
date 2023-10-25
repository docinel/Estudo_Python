import pyautogui as pa
import timeit
import openpyxl as xl


start_time = timeit.default_timer()

wb = xl.load_workbook(r'Z:\011-TI\017-BI_TESTES\BI_BASE_RELATORIOS\IMPL_SALDO.xlsx')
ws = wb['IMPL_SALDO']

qtd = int(pa.prompt(text='Digite a quantidade de Implantações:',
                    title='QTD LINHAS', default=''))
#  ABRE O MENTOR
pa.hotkey('win', 'd')
pa.sleep(1)
pa.getWindowsWithTitle('Touch Comp ERP')[0].maximize()

# NOVO ITEM
pa.click(55, 296)
meu_dic = {}
pa.sleep(5)
for cont in range(2, qtd):
    meu_dic[f'A{cont}'] = ws[f'A{cont}'].value
    meu_dic[f'B{cont}'] = ws[f'B{cont}'].value

for cont in range(2, qtd):
    pa.write(str(meu_dic[f'A{cont}']))
    pa.sleep(1)
    pa.press('enter')
    pa.sleep(1)
    pa.typewrite('3')
    pa.sleep(1)
    pa.press('enter')
    pa.write(str(meu_dic[f'B{cont}']))
    pa.press('enter')
    pa.write('1')
    pa.press('enter', presses=4, interval=0.3)

# Finalizando e Salvando
pa.click(196, 130)

print((timeit.default_timer() - start_time)/60)
