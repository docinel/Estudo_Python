import openpyxl as xl

wb = xl.load_workbook(r'sources/excel/calculo.xlsx')
ws = wb.active

lista = []
for cont in range(2, round(ws.max_row) + 1):
    lista.append(ws[f'A{cont}'].value)

string_formatada = ', '.join(f"'{item}'" for item in lista)

print(lista, str(len(lista)))
print(string_formatada)
wb.close()
