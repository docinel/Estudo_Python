import openpyxl as xl

wb = xl.load_workbook(r'sources/excel/calculo.xlsx')
ws = wb.active

teste=0
while teste < 20:
    teste1 = teste
    teste +=1
# print(wb.sheetnames)
# for row in ws.iter_rows(min_row=2, values_only=True):
#     for cell in row:
#         print(cell)
#for row in ws.iter_rows(min_row=2,values_only=True):
cont = teste1
qtd_max_cel = int(input('Qtd linhas: '))
while cont < qtd_max_cel:
    celula = ws[f'A{teste1}']
    print(celula.value)
    cont +=1
    
# celula = ws['A2']
# celula2 = ws['B2']

# print(celula.value)
# print(celula2.value)