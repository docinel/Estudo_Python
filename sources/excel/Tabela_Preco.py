import pyautogui as pa
import timeit
import openpyxl as xl

pa.FAILSAFE = True

start_time = timeit.default_timer()

wb = xl.load_workbook(r'Z:\017-CUSTOS\002_CONTROLES\CONTROLE_DE_CUSTOS\CONTROLE_CUSTOS.xlsx')
ws = wb['CUSTOS_TOTAL']

pa.hotkey('win', 'd')
qtd = int(pa.prompt(text='Digite a quantidade de Inserções:',
                    title='TABELA DE PREÇOS', default='')) + 2
#  ABRE O MENTOR
pa.hotkey('win', 'd')
pa.sleep(1)
pa.getWindowsWithTitle('Touch Comp ERP')[0].maximize()

meu_dic = {}
for cont in range(2, qtd):
    # ID DO PRODUTO4
    meu_dic[f'D{cont}'] = ws[f'D{cont}'].value
    # VALOR DO CUSTO
    meu_dic[f'E{cont}'] = ws[f'E{cont}'].value
    # VALOR DO PRODUTO
    meu_dic[f'C{cont}'] = ws[f'C{cont}'].value


for cont in range(2, qtd):
    # NOVO ITEM
    pa.click(406, 327)
    pa.sleep(1)
    pa.press('a')

    # INSERIR ID PRODUTO
    pa.sleep(2)
    pa.press('tab', presses=2, interval=0.4)
    pa.write(str(meu_dic[f'D{cont}']), interval=0.3)
    pa.press('enter')
    pa.sleep(5)
    pa.press('enter')

    # INSERIR VALOR DO CUSTO
    pa.sleep(5)
    pa.click(643, 402)
    pa.sleep(1)
    pa.write(str(meu_dic[f'E{cont}']).replace('.', ','), interval=0.3)

    # INSERIR VALOR DO PRODUTO
    pa.sleep(1)
    pa.press('tab', presses=2, interval=1)
    pa.sleep(1)
    pa.write(str(meu_dic[f'C{cont}']).replace('.', ','), interval=0.3)
    pa.press('enter')

pa.sleep(2)
#  Finalizando e Salvando
pa.click(196, 130)

# Selecionar a linha a ser apagada
ws.delete_rows(2, qtd-2)
wb.save(r'Z:\017-CUSTOS\002_CONTROLES\CONTROLE_DE_CUSTOS\CONTROLE_CUSTOS.xlsx')


end_time = timeit.default_timer() - start_time

result = round(end_time / 60, 2)

if result >= 1:
    pa.alert(text='O processo levou {} minuto(s)'.format(result), title='SUCESSO', button='OK')
else:
    pa.alert(text='O processo levou {} segundo(s)'.format(result), title='SUCESSO', button='OK')
