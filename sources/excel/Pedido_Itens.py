import pyautogui as pa
import timeit
import openpyxl as xl

pa.FAILSAFE = True

start_time = timeit.default_timer()

wb = xl.load_workbook(r"C:\Users\narobo\Desktop\LEILAO\LEILOES.xlsx")
ws = wb['DADOS']

qtd = int(pa.prompt(text='Digite a quantidade de Produtos:',
                    title='QTD PRODUTOS', default='')) + 2
#  ABRE O MENTOR
pa.hotkey('win', 'd')
pa.sleep(0.6)
pa.getWindowsWithTitle('Touch Comp ERP')[0].maximize()

# NOVO ITEM
pa.click(173, 192)
meu_dic = {}
pa.sleep(3)
for cont in range(2, qtd):
    meu_dic[f'A{cont}'] = ws[f'A{cont}'].value
    meu_dic[f'B{cont}'] = ws[f'B{cont}'].value
    meu_dic[f'C{cont}'] = ws[f'C{cont}'].value

for cont in range(2, qtd):
    pa.click(172, 198)
    pa.sleep(1)
    pa.write(str(meu_dic[f'A{cont}']))
    pa.press('tab')
    pa.sleep(5)
    pa.write(str(meu_dic[f'B{cont}']))
    pa.press('enter')
    pa.sleep(1.5)
    pa.write(str(meu_dic[f'C{cont}']))
    pa.sleep(1.5)
    pa.press('enter', presses=4, interval=0.3)

    pa.click(640, 445)
    pa.click(171, 193)
    pa.sleep(1)
    pa.click(213, 325)
    pa.sleep(1)
# Finalizando e Salvando
pa.click(200, 132)
wb.close()

print((timeit.default_timer() - start_time)/60)
