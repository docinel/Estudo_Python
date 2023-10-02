import pyautogui as pa
from openpyxl import Workbook, load_workbook
import clipboard
import timeit

start_time = timeit.default_timer()

# DEFINIR A QTD DE EMAILS
qtd = int(pa.prompt(text='Digite a quantidade de Implantações:', title='QTD EMAILS' , default=''))
wb = load_workbook(r'C:\workspaces\Estudo_Python\sources\outlook\spam.xlsx')
ws = wb.active
pa.hotkey('win', 'r'), pa.sleep(1)

pa.typewrite('outlook'), pa.sleep(1), pa.press('enter')
pa.sleep(5)
pa.click(110, 220), pa.sleep(1)
contador = 0
while contador < qtd :    
    pa.click(404, 185), pa.sleep(0.5)
    pa.rightClick(676, 126), pa.sleep(2)
    pa.typewrite('p')
    copia = clipboard.paste()
    pa.click(550, 193)
    contador +=1
    ws[f'A{contador}'].value = copia


wb.save('sources\outlook\spam.xlsx')
print((timeit.default_timer() - start_time)/60)