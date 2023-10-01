from openpyxl import Workbook, load_workbook

wb = load_workbook('sources\excel_2\pasta.xlsx')
wb.create_sheet('Dados')

# ws = wb.active
# ws['A1'].value = 'Nomes'


wb.save('sources\excel_2\pasta.xlsx')
# print(ws['A1'].value)
print(wb.sheetnames)